import openpyxl


class HomePageData:

    test_Homepage_data =[{"Firstname":"vijay","Lastname":"kumar","Gender":"Male"},{"Firstname":"vk","Lastname":"vijay","Gender":"Male"}]

    @staticmethod
    def getTestData(test_case_name):
        book = openpyxl.load_workbook("C:\\Users\\vijay\\OneDrive\\Documents\\pythonExcelDemo.xlsx")
        sheet = book.active
        Dict={}
        for i in range(1, sheet.max_row + 1):   # to get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):     # to get columns
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [Dict]